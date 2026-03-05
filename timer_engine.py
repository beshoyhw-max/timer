"""Meeting Speaker Timer — State Machine, History, Tiered Cost & JSON Config."""

import math
import sys
import time
import csv
import io
import os
import json
from datetime import datetime
from enum import Enum
from dataclasses import dataclass, field
from typing import Optional

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ── Writable data directory (works as .py script or packaged .exe) ──
if getattr(sys, 'frozen', False):
    _DATA_DIR = os.path.dirname(sys.executable)
else:
    _DATA_DIR = os.path.dirname(os.path.abspath(__file__))

CONFIG_PATH = os.path.join(_DATA_DIR, "config.json")


class Phase(str, Enum):
    IDLE = "idle"
    COUNTDOWN = "countdown"
    WARNING = "warning"
    OVERTIME = "overtime"
    COLLECTING = "collecting"
    THANKYOU = "thankyou"
    PAUSED = "paused"


@dataclass
class CostTier:
    threshold_mins: float  # overtime minutes where this tier starts
    rate_amount: float     # ¥ per tick (or per minute if unit="min")
    rate_interval: int     # seconds per tick (ignored if unit="min")
    unit: str = "sec"      # "sec" or "min"

    def to_dict(self):
        return {
            "threshold_mins": self.threshold_mins,
            "rate_amount": self.rate_amount,
            "rate_interval": self.rate_interval,
            "unit": self.unit,
        }

    @staticmethod
    def from_dict(d):
        return CostTier(
            threshold_mins=float(d.get("threshold_mins", 0)),
            rate_amount=float(d.get("rate_amount", 5)),
            rate_interval=int(d.get("rate_interval", 5)),
            unit=str(d.get("unit", "sec")),
        )


@dataclass
class SpeakerRecord:
    name: str
    allocated_seconds: int
    tiers: list = field(default_factory=list)
    max_cost: float = 2500.0
    actual_seconds: float = 0
    overtime_seconds: float = 0
    cost: float = 0
    time_added: int = 0
    start_dt: str = ""
    end_dt: str = ""


class TimerEngine:
    WARNING_THRESHOLD = 60  # seconds before end for visual warning

    def __init__(self):
        self.phase = Phase.IDLE
        self.speaker_name = ""
        self.allocated_seconds = 0
        self.time_added = 0
        self.start_time: Optional[float] = None
        self.overtime_start: Optional[float] = None
        self.paused_at: Optional[float] = None
        self.pre_pause_phase: Optional[Phase] = None
        self.history: list[SpeakerRecord] = []
        self._current_record: Optional[SpeakerRecord] = None

        # Tiered cost config
        self.tiers: list[CostTier] = [CostTier(0, 5.0, 5, "sec")]
        self.max_cost: float = 2500.0

        # Alarm thresholds (seconds remaining)
        self.alarm_threshold_1: float = 300  # 5 minutes
        self.alarm_threshold_2: float = 60   # 1 minute
        self.alarm_3_interval: float = 60    # Alarm 3: repeat every X seconds during overtime
        self.alarms_enabled: bool = True     # Master toggle for sounds
        self._alarm_1_fired: bool = False
        self._alarm_2_fired: bool = False
        self._alarm_3_last_fired: float = 0  # timestamp of last alarm 3 firing
        self.alarm_sounds: dict = {"1": "1.wav", "2": "2.wav", "3": "3.wav"}
        self.alarm_durations: dict = {"1": 0.0, "2": 0.0, "3": 0.0}

        # Appearance config
        self.appearance: dict = {
            "background_image": "premium_bg.png",
            "color_countdown": "#2dd4bf",
            "color_warning": "#fbbf24",
            "color_overtime": "#ef4444",
            "color_thankyou": "#f1c40f",
            "color_cost": "#ff7979",
            "color_quote": "#f1c40f",
            "font_family": "JetBrains Mono",
            "text_on_time": "准时完成!",
            "text_overtime": "议题超时，请赞助",
        }

        # Default language
        self.language: str = "en"

        # Quote & Speaker list
        self.quote: str = ""
        self.speakers: list[dict] = []  # [{name, time_secs, words}]
        self._next_speaker_index: int = 0

        # Export folder (writable, next to EXE)
        self._export_dir = os.path.join(_DATA_DIR, "exports")

        # Load saved config
        self._load_config()

    # ── Config Persistence ────────────────────────────────────────

    def _load_config(self):
        if not os.path.exists(CONFIG_PATH):
            return
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
            if "tiers" in data:
                self.tiers = [CostTier.from_dict(t) for t in data["tiers"]]
            if "max_cost" in data:
                self.max_cost = float(data["max_cost"])
            if "alarm_threshold_1" in data:
                self.alarm_threshold_1 = float(data["alarm_threshold_1"])
            if "alarm_threshold_2" in data:
                self.alarm_threshold_2 = float(data["alarm_threshold_2"])
            if "alarm_3_interval" in data:
                self.alarm_3_interval = float(data["alarm_3_interval"])
            if "alarms_enabled" in data:
                self.alarms_enabled = bool(data["alarms_enabled"])
            if "quote" in data:
                self.quote = str(data["quote"])
            if "language" in data:
                self.language = str(data["language"])
            if "speakers" in data:
                self.speakers = list(data["speakers"])
            if "appearance" in data:
                self.appearance.update(data["appearance"])
            if "alarm_sounds" in data:
                self.alarm_sounds.update(data["alarm_sounds"])
            if "alarm_durations" in data:
                # convert dictionary to floats just in case
                self.alarm_durations.update({k: float(v) for k, v in data["alarm_durations"].items()})
        except Exception as e:
            print(f"[timer] Failed to load config: {e}")

    def save_config(self, tiers_data: list, max_cost: float,
                    alarm_1: float, alarm_2: float, alarms_enabled: bool = True,
                    alarm_3_interval: float = 60,
                    alarm_sounds: dict = None, alarm_durations: dict = None):
        """Save tier config + alarm thresholds to JSON and update in-memory."""
        self.tiers = [CostTier.from_dict(t) for t in tiers_data]
        self.max_cost = max_cost
        self.alarm_threshold_1 = alarm_1
        self.alarm_threshold_2 = alarm_2
        self.alarm_3_interval = alarm_3_interval
        self.alarms_enabled = alarms_enabled
        if alarm_sounds:
            self.alarm_sounds.update(alarm_sounds)
        if alarm_durations:
            self.alarm_durations.update({k: float(v) for k, v in alarm_durations.items()})

        data = {
            "tiers": [t.to_dict() for t in self.tiers],
            "max_cost": self.max_cost,
            "alarm_threshold_1": self.alarm_threshold_1,
            "alarm_threshold_2": self.alarm_threshold_2,
            "alarm_3_interval": self.alarm_3_interval,
            "alarms_enabled": self.alarms_enabled,
            "alarm_sounds": dict(self.alarm_sounds),
            "alarm_durations": dict(self.alarm_durations),
            "language": self.language,
            "quote": self.quote,
            "speakers": self.speakers,
            "appearance": dict(self.appearance),
        }
        try:
            with open(CONFIG_PATH, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"[timer] Failed to save config: {e}")

    def get_config(self) -> dict:
        return {
            "tiers": [t.to_dict() for t in self.tiers],
            "max_cost": self.max_cost,
            "alarm_threshold_1": self.alarm_threshold_1,
            "alarm_threshold_2": self.alarm_threshold_2,
            "alarm_3_interval": self.alarm_3_interval,
            "alarms_enabled": self.alarms_enabled,
            "alarm_sounds": dict(self.alarm_sounds),
            "alarm_durations": dict(self.alarm_durations),
            "language": self.language,
            "quote": self.quote,
            "speakers": self.speakers,
            "appearance": dict(self.appearance),
        }

    # ── Quote & Language ──────────────────────────────────────────

    def set_quote(self, text: str):
        self.quote = text
        self._persist_config()

    def get_language(self) -> str:
        return getattr(self, "language", "en")

    def set_language(self, lang: str):
        self.language = lang
        self._persist_config()

    # ── Appearance ─────────────────────────────────────────────────

    def get_appearance(self) -> dict:
        return dict(self.appearance)

    def set_appearance(self, data: dict):
        self.appearance.update(data)
        self._persist_config()

    # ── Speakers List ─────────────────────────────────────────────

    def get_speakers(self) -> list:
        return self.speakers

    def set_speakers(self, speakers_data: list):
        # Only reset index if the list size changed (import/add/remove)
        if len(speakers_data) != len(self.speakers):
            self._next_speaker_index = 0
        self.speakers = speakers_data
        self._persist_config()

    def reset_speaker_list(self):
        """Reset speaker list index to start from the beginning."""
        self._next_speaker_index = 0


    def load_next_speaker(self) -> dict:
        """Return the next speaker from the list and advance the index."""
        if not self.speakers:
            return {}
        if self._next_speaker_index >= len(self.speakers):
            self._next_speaker_index = 0
        
        return self.load_specific_speaker(self._next_speaker_index)

    def load_specific_speaker(self, index: int) -> dict:
        """Load a specific speaker by index and set them as the current/next speaker."""
        if not self.speakers or index < 0 or index >= len(self.speakers):
            return {}
            
        self._next_speaker_index = index
        speaker = dict(self.speakers[index])
        speaker["_index"] = index
        
        # Advance index for the next "Next Speaker" click
        self._next_speaker_index += 1
        
        # Use speaker's words as the animation quote
        if speaker.get("words"):
            self.quote = speaker["words"]
            
        return speaker

    def load_previous_speaker(self) -> dict:
        """Go back to the previous speaker. Safety net for accidental double-click on Next."""
        if not self.speakers:
            return {}
        # _next_speaker_index points to the NEXT speaker to load.
        # Current loaded speaker is at _next_speaker_index - 1.
        # Previous speaker is at _next_speaker_index - 2.
        prev_idx = self._next_speaker_index - 2
        if prev_idx < 0:
            prev_idx = len(self.speakers) - 1  # wrap around
        return self.load_specific_speaker(prev_idx)

    def get_upcoming_speakers(self, count: int = 5) -> list:
        """Return upcoming speakers from the list without advancing the pointer."""
        if not self.speakers:
            return []
        result = []
        idx = self._next_speaker_index
        for _ in range(min(count, len(self.speakers))):
            if idx >= len(self.speakers):
                idx = 0
            speaker = dict(self.speakers[idx])
            speaker["_index"] = idx
            result.append(speaker)
            idx += 1
        return result

    def import_speakers_from_excel(self, filepath: str) -> list:
        """Read speakers from an Excel file. Columns: Speaker Name, time in secs, words."""
        if not HAS_OPENPYXL:
            print("[timer] openpyxl not installed")
            return []

        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            speakers = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) < 2:
                    continue
                name = str(row[0] or "").strip()
                if not name:
                    continue
                val = row[1]
                # Check for "MM:SS" format (string)
                if isinstance(val, str) and ":" in val:
                    parts = val.split(":")
                    if len(parts) == 2:
                        try:
                            m = float(parts[0])
                            s = float(parts[1])
                            time_secs = int(m * 60 + s)
                        except ValueError:
                            time_secs = 0
                    else:
                        time_secs = 0
                # Handle datetime.time objects from openpyxl
                # (e.g. Excel cells formatted as Time → openpyxl returns time object)
                elif hasattr(val, 'hour') and hasattr(val, 'minute') and hasattr(val, 'second'):
                    try:
                        time_secs = int(val.hour * 3600 + val.minute * 60 + val.second)
                    except AttributeError:
                         time_secs = 0
                else:
                    try:
                        time_secs = int(float(val or 0))
                    except (ValueError, TypeError):
                        time_secs = 0
                words = str(row[2] or "").strip() if len(row) > 2 else ""
                speakers.append({
                    "name": name,
                    "time_secs": time_secs,
                    "words": words,
                })
            wb.close()
            self.speakers = speakers
            self._next_speaker_index = 0
            self._persist_config()
            return speakers
        except Exception as e:
            print(f"[timer] Failed to import Excel: {e}")
            return []

    def generate_excel_template(self, filepath: str):
        """Generate a sample Excel template for speaker list import."""
        if not HAS_OPENPYXL:
            print("[timer] openpyxl not installed")
            return False

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Speaker List Template"

            # Headers
            headers = ["Speaker/Topic Name", "Duration (MM:SS or seconds)", "Quote/Words (Optional)"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[chr(64 + col_idx)].width = 25

            # Sample Data
            samples = [
                ["Alice", "05:00", "Focus on the goal."],
                ["Bob", "10:00", "Time is money."],
                ["Charlie", "180", "Keep it simple."]
            ]
            for row_idx, sample in enumerate(samples, 2):
                for col_idx, value in enumerate(sample, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            wb.save(filepath)
            return True
        except Exception as e:
            print(f"[timer] Failed to generate Excel template: {e}")
            return False

    def _persist_config(self):
        """Helper to persist current config to disk."""
        data = {
            "tiers": [t.to_dict() for t in self.tiers],
            "max_cost": self.max_cost,
            "alarm_threshold_1": self.alarm_threshold_1,
            "alarm_threshold_2": self.alarm_threshold_2,
            "alarm_3_interval": self.alarm_3_interval,
            "alarms_enabled": self.alarms_enabled,
            "alarm_sounds": dict(self.alarm_sounds),
            "alarm_durations": dict(self.alarm_durations),
            "language": self.language,
            "quote": self.quote,
            "speakers": self.speakers,
            "appearance": dict(self.appearance),
        }
        try:
            with open(CONFIG_PATH, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"[timer] Failed to save config: {e}")

    # ── Timer Control ─────────────────────────────────────────────

    def configure(self, speaker: str, minutes: float, seconds: float):
        self.speaker_name = speaker
        self.allocated_seconds = minutes * 60 + seconds

    def start(self):
        if self.phase != Phase.IDLE:
            return
        self.phase = Phase.COUNTDOWN
        self.start_time = time.time()
        self.overtime_start = None
        self.time_added = 0
        self.paused_at = None
        self.pre_pause_phase = None
        self._alarm_1_fired = False
        self._alarm_2_fired = False
        self._alarm_3_last_fired = 0
        self._current_record = SpeakerRecord(
            name=self.speaker_name,
            allocated_seconds=self.allocated_seconds,
            tiers=[t.to_dict() for t in self.tiers],
            max_cost=self.max_cost,
            start_dt=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        )

    def pause(self):
        """Pause the timer, recording the time."""
        if self.phase in (Phase.COUNTDOWN, Phase.WARNING, Phase.OVERTIME):
            self.pre_pause_phase = self.phase
            self.phase = Phase.PAUSED
            self.paused_at = time.time()

    def resume(self):
        """Resume the timer, adjusting start_time by the paused duration."""
        if self.phase == Phase.PAUSED and self.pre_pause_phase:
            pause_duration = time.time() - (self.paused_at or time.time())
            if self.start_time:
                self.start_time += pause_duration
            if self.overtime_start:
                self.overtime_start += pause_duration
            if self._alarm_3_last_fired > 0:
                self._alarm_3_last_fired += pause_duration
            
            self.phase = self.pre_pause_phase
            self.paused_at = None
            self.pre_pause_phase = None

    def stop(self):
        if self.phase in (Phase.IDLE, Phase.COLLECTING, Phase.THANKYOU):
            return
        state = self.get_state()
        self.phase = Phase.COLLECTING
        if self._current_record:
            self._current_record.actual_seconds = state["elapsed"]
            self._current_record.overtime_seconds = state["overtime_seconds"]
            
            # Final Cost Calculation with "Finish a Minute" logic
            # If the FINAL active tier is 'min' unit, we round the total overtime duration 
            # up to the next full minute for billing purposes.
            final_overtime_calc = state["overtime_seconds"]
            
            # Determine applicable tier at the end
            current_tier = None
            if self.tiers and state["overtime_seconds"] > 0:
                sorted_tiers = sorted(self.tiers, key=lambda t: t.threshold_mins)
                for t in sorted_tiers:
                    if state["overtime_seconds"] / 60 >= t.threshold_mins:
                        current_tier = t
                    else:
                        break
            
            if current_tier and current_tier.unit == "min":
                # Round up total overtime to next minute
                # e.g. 1m 51s (111s) -> 120s
                if final_overtime_calc > 0:
                    minutes = math.ceil(final_overtime_calc / 60.0)
                    final_overtime_calc = minutes * 60.0
            
            # Recalculate cost with potentially rounded time
            self._current_record.cost = self._compute_cost(final_overtime_calc)
            
            self._current_record.time_added = self.time_added
            self._current_record.end_dt = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.history.append(self._current_record)
            self._export_to_excel(self._current_record)

    def finish_animation(self):
        if self.phase != Phase.COLLECTING:
            return
        self.phase = Phase.THANKYOU

    def next_speaker(self):
        self.phase = Phase.IDLE
        self.speaker_name = ""
        self.start_time = None
        self.overtime_start = None
        self.time_added = 0
        self._current_record = None
        self._alarm_1_fired = False
        self._alarm_2_fired = False

    def add_time(self, seconds: float):
        """Add time to running timer. Can revert WARNING→COUNTDOWN."""
        if self.phase in (Phase.IDLE, Phase.COLLECTING, Phase.THANKYOU):
            return
        self.allocated_seconds += seconds
        self.time_added += seconds

        # Recalculate: if we're in warning/overtime but now have enough time,
        # revert phase.
        
        # Calculate effective NOW
        now = time.time()
        if self.phase == Phase.PAUSED and self.paused_at:
            now = self.paused_at

        elapsed = (now - self.start_time) if self.start_time else 0
        remaining = self.allocated_seconds - elapsed

        check_phase = self.pre_pause_phase if self.phase == Phase.PAUSED else self.phase

        if remaining > self.WARNING_THRESHOLD:
            if check_phase in (Phase.WARNING, Phase.OVERTIME):
                new_phase = Phase.COUNTDOWN
                self.overtime_start = None
                if self.phase == Phase.PAUSED:
                    self.pre_pause_phase = new_phase
                else:
                    self.phase = new_phase
                    
        elif remaining > 0:
            if check_phase == Phase.OVERTIME:
                new_phase = Phase.WARNING
                self.overtime_start = None
                if self.phase == Phase.PAUSED:
                    self.pre_pause_phase = new_phase
                else:
                    self.phase = new_phase
        else:
            # We are still in OVERTIME (or just pushed into it), but we added time.
            # We must shift overtime_start forward by the added seconds so the clock updates.
            if check_phase == Phase.OVERTIME and self.overtime_start:
                self.overtime_start += seconds

        # Reset alarm flags if remaining goes back above thresholds
        if remaining > self.alarm_threshold_1:
            self._alarm_1_fired = False
        if remaining > self.alarm_threshold_2:
            self._alarm_2_fired = False

    def subtract_time(self, seconds: float):
        """Subtract time from running timer. Can push COUNTDOWN→WARNING→OVERTIME."""
        if self.phase in (Phase.IDLE, Phase.COLLECTING, Phase.THANKYOU):
            return
        self.allocated_seconds -= seconds
        self.time_added -= seconds

        # Recalculate remaining
        now = time.time()
        if self.phase == Phase.PAUSED and self.paused_at:
            now = self.paused_at

        elapsed = (now - self.start_time) if self.start_time else 0
        remaining = self.allocated_seconds - elapsed

        check_phase = self.pre_pause_phase if self.phase == Phase.PAUSED else self.phase

        if remaining <= 0:
            if check_phase != Phase.OVERTIME:
                new_phase = Phase.OVERTIME
                self.overtime_start = now - abs(remaining)
                self._alarm_3_last_fired = now
                if self.phase == Phase.PAUSED:
                    self.pre_pause_phase = new_phase
                else:
                    self.phase = new_phase
            else:
                # We are ALREADY in overtime, and we subtracted more time.
                # Overtime INCREASES. Example: we were 3 mins in overtime (now - start = 3).
                # We subtract 5 mins of allocation. We are now conceptually 8 mins in overtime.
                # Thus, overtime_start must move BACKWARDS by `seconds` so `now - (start - secs)` is larger.
                if self.overtime_start:
                    self.overtime_start -= seconds
                    
        elif remaining <= self.WARNING_THRESHOLD:
            if check_phase == Phase.COUNTDOWN:
                new_phase = Phase.WARNING
                if self.phase == Phase.PAUSED:
                    self.pre_pause_phase = new_phase
                else:
                    self.phase = new_phase

    # ── Tiered Cost Calculation ───────────────────────────────────

    def _compute_cost(self, overtime_seconds: float) -> float:
        """
        Calculate cost based on tiers.
        If unit="min", calculation is PROPORTIONAL (per second) here.
        Rounding up happens only at the 'stop' event if the final tier is 'min'.
        """
        if overtime_seconds <= 0 or not self.tiers:
            return 0.0

        # Sort tiers by threshold
        sorted_tiers = sorted(self.tiers, key=lambda t: t.threshold_mins)
        total_cost = 0.0

        for i, tier in enumerate(sorted_tiers):
            tier_start_secs = tier.threshold_mins * 60
            if i + 1 < len(sorted_tiers):
                tier_end_secs = sorted_tiers[i + 1].threshold_mins * 60
            else:
                tier_end_secs = float("inf")

            if overtime_seconds <= tier_start_secs:
                break

            # Effective duration in this tier
            tier_active_duration = min(overtime_seconds, tier_end_secs) - tier_start_secs
            
            if tier_active_duration > 0:
                if tier.unit == "min":
                    # Proportional calculation: (duration_min) * (rate_per_min)
                    minutes = tier_active_duration / 60.0
                    total_cost += minutes * tier.rate_amount
                else:
                    # Default: Proportional / Interval based (e.g. every 5 secs)
                    if tier.rate_interval > 0:
                        ticks = tier_active_duration / tier.rate_interval
                        total_cost += ticks * tier.rate_amount

        return min(total_cost, self.max_cost)

    # ── Excel Auto-Export ─────────────────────────────────────────

    def _export_to_excel(self, record: SpeakerRecord):
        if not HAS_OPENPYXL:
            print("[timer] openpyxl not installed — skipping Excel export")
            return

        os.makedirs(self._export_dir, exist_ok=True)

        filename = datetime.now().strftime("%Y%m%d") + ".xlsx"
        filepath = os.path.join(self._export_dir, filename)

        headers = [
            "Speaker",
            "Allocated Time",
            "Time Added",
            "Cost Tiers",
            "Actual Duration",
            "Start Time",
            "End Time",
            "Amount Due (￥)",
        ]

        am, as_ = divmod(int(record.actual_seconds), 60)
        m, s = divmod(int(record.allocated_seconds), 60)

        # Format tiers for display
        tier_strs = []
        for t in (record.tiers if record.tiers else []):
            if isinstance(t, dict):
                tier_strs.append(
                    f">{t.get('threshold_mins', 0)}min: "
                    f"￥{t.get('rate_amount', 0)}/{t.get('rate_interval', 5)}s"
                )
        tier_display = " | ".join(tier_strs) if tier_strs else "N/A"

        added_m, added_s = divmod(int(record.time_added), 60)

        row = [
            record.name,
            f"{m:02d}:{s:02d}",
            f"+{added_m:02d}:{added_s:02d}" if record.time_added else "—",
            tier_display,
            f"{am:02d}:{as_:02d}",
            record.start_dt,
            record.end_dt,
            f"￥{record.cost:.2f}",
        ]

        if os.path.exists(filepath):
            wb = load_workbook(filepath)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Meeting Log"

            header_font = Font(name="Inter", bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                bottom=Side(style="thin", color="333333")
            )

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            widths = [18, 16, 14, 30, 16, 22, 22, 18]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[chr(64 + i)].width = w

        ws.append(row)
        wb.save(filepath)

    # ── State ─────────────────────────────────────────────────────

    def get_state(self) -> dict:
        now = time.time()
        if self.phase == Phase.PAUSED and self.paused_at:
            now = self.paused_at
            
        elapsed = (now - self.start_time) if self.start_time else 0
        remaining = max(0, self.allocated_seconds - elapsed)

        # Auto-transition countdown → warning → overtime
        if self.phase == Phase.COUNTDOWN and remaining <= self.WARNING_THRESHOLD:
            self.phase = Phase.WARNING
        if self.phase == Phase.WARNING and remaining <= 0:
            self.phase = Phase.OVERTIME
            self.overtime_start = now - (elapsed - self.allocated_seconds)
            self._alarm_3_last_fired = now  # Start interval from overtime begin

        # Alarm triggers (one-shot)
        # Skip alarm if threshold >= allocated time (entire meeting is shorter
        # than the warning window, so the alarm would fire immediately).
        alarm_1_trigger = False
        alarm_2_trigger = False
        alarm_3_trigger = False
        # Anticipation offset of 1.5 seconds to combat audio latency, winsound/Windows sleep padding, 
        # and visual clock rounding (since 01:00 appears precisely at 60.9 seconds remaining).
        # This makes the alarm audible exactly when the visual clock reaches the target.
        alarm_offset = 1
        if self.phase in (Phase.COUNTDOWN, Phase.WARNING, Phase.OVERTIME):
            if (self.alarm_threshold_1 < self.allocated_seconds
                    and not self._alarm_1_fired
                    and remaining <= self.alarm_threshold_1 + alarm_offset):
                self._alarm_1_fired = True
                if self.alarms_enabled and (self.phase != Phase.OVERTIME or self.alarm_threshold_1 == 0):
                    alarm_1_trigger = True
            if (self.alarm_threshold_2 < self.allocated_seconds
                    and not self._alarm_2_fired
                    and remaining <= self.alarm_threshold_2 + alarm_offset):
                self._alarm_2_fired = True
                if self.alarms_enabled and (self.phase != Phase.OVERTIME or self.alarm_threshold_2 == 0):
                    alarm_2_trigger = True

        # Alarm 3: recurring alarm during overtime
        if self.phase == Phase.OVERTIME and self.alarm_3_interval > 0:
            if now - self._alarm_3_last_fired >= self.alarm_3_interval:
                if self.alarms_enabled:
                    alarm_3_trigger = True
                self._alarm_3_last_fired = now

        overtime_seconds = 0.0
        cost = 0.0
        is_overtime = (self.phase == Phase.OVERTIME) or (self.phase == Phase.PAUSED and self.pre_pause_phase == Phase.OVERTIME)
        if is_overtime and self.overtime_start:
            overtime_seconds = now - self.overtime_start
            cost = self._compute_cost(overtime_seconds)
        elif self.phase in (Phase.COLLECTING, Phase.THANKYOU) and self._current_record:
            overtime_seconds = self._current_record.overtime_seconds
            cost = self._current_record.cost

        minutes_r, seconds_r = divmod(int(remaining), 60)

        return {
            "phase": self.phase.value,
            "speaker": self.speaker_name,
            "allocated": self.allocated_seconds,
            "elapsed": round(elapsed, 1),
            "remaining": round(remaining, 1),
            "remaining_display": f"{minutes_r:02d}:{seconds_r:02d}",
            "overtime_seconds": round(overtime_seconds, 1),
            "cost": round(cost, 2),
            "progress": min(1.0, elapsed / self.allocated_seconds) if self.allocated_seconds > 0 else 0,
            "time_added": self.time_added,
            "alarm_1": alarm_1_trigger,
            "alarm_2": alarm_2_trigger,
            "alarm_3": alarm_3_trigger,
            "alarms_enabled": self.alarms_enabled,
            "list_index": self._next_speaker_index,
            "quote": self.quote,
            "appearance": dict(self.appearance),
        }

    def get_history(self) -> list[dict]:
        records = []
        for r in self.history:
            m, s = divmod(int(r.allocated_seconds), 60)
            am, as_ = divmod(int(r.actual_seconds), 60)
            records.append({
                "name": r.name,
                "allocated": f"{m:02d}:{s:02d}",
                "actual": f"{am:02d}:{as_:02d}",
                "overtime": round(r.overtime_seconds, 1),
                "cost": round(r.cost, 2),
                "time_added": r.time_added,
            })
        return records

    def export_csv(self) -> str:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Speaker", "Allocated", "Actual", "Time Added (s)",
                         "Overtime (s)", "Cost (￥)"])
        for r in self.history:
            m, s = divmod(int(r.allocated_seconds), 60)
            am, as_ = divmod(int(r.actual_seconds), 60)
            writer.writerow([
                r.name,
                f"{m:02d}:{s:02d}",
                f"{am:02d}:{as_:02d}",
                r.time_added,
                round(r.overtime_seconds, 1),
                round(r.cost, 2),
            ])
        return output.getvalue()
